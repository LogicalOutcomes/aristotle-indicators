from __future__ import unicode_literals
import datetime
from aristotle_mdr import models
from aristotle_mdr.contrib.identifiers import models as MDR_ID
from aristotle_mdr.contrib.slots.models import Slot
from comet import models as comet
from indicators import models as lo_models
from mallard_qr import models as mallard_qr
from openpyxl import load_workbook
from .utils import lb_2_p, get_col


class IndicatorImporter(object):
    """Import indicators from spreadsheet
    """
    DEFAULT_AUTHORITY_NAME = 'Logical Outcomes'
    DEFAULT_AUTHORITY_PREFIX = 'lo'
    DEFAULT_REGISTRATION_DATE = datetime.date(2009, 4, 28)
    DEFAULT_DEFINITION = 'Imported from the Indicator Reference Sheet'
    # data file sheets
    SHEET_GOALS = 'Sustainable Development Goals'
    SHEET_INSTRUMENTS = 'Instruments'
    SHEET_VALUE_DOMAIN = 'Value Domain'
    SHEET_INDICATORS = 'Indicators'
    SHEET_DATA_ELEMENTS = 'Data Elements'

    def __init__(self, data_file):
        self.wb = load_workbook(data_file, read_only=True)

    def process(self):
        self.process_authorities()
        self.process_goals()
        self.process_instruments()
        self.process_value_domains()
        self.process_indicators()
        self.process_data_elements()

    def process_authorities(self):
        self.authority, created = models.RegistrationAuthority.objects.get_or_create(
            name=self.DEFAULT_AUTHORITY_NAME
        )
        self.authority_org = models.Organization.objects.get(
            pk=self.authority.pk
        )
        self.authority_namespace, c = MDR_ID.Namespace.objects.get_or_create(
            naming_authority=self.authority_org,
            shorthand_prefix=self.DEFAULT_AUTHORITY_PREFIX
        )

    def process_goals(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_GOALS)
        for row in sheet.iter_rows(row_offset=1):
            if row[0].value is None:
                continue

            goal, c = lo_models.Goal.objects.get_or_create(
                short_name=get_col(row, 'A').value,
                name=get_col(row, 'B').value,
                definition=get_col(row, 'C').value,
                origin_URI=get_col(row, 'D').value,
            )
            self.register(goal)

    def process_instruments(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_INSTRUMENTS)
        for row in sheet.iter_rows(row_offset=1):
            if row[2].value is None:
                continue
            name = get_col(row, 'A').value

            defaults = {
                "definition": lb_2_p(get_col(row, 'E').value or self.DEFAULT_DEFINITION, sep="\n"),
                "terms_of_use": get_col(row, 'D').value or "",
                "where_to_get": get_col(row, 'C').value or "",
                "references": get_col(row, 'B').value or "",
            }
            inst, c = lo_models.Instrument.objects.update_or_create(
                name=name,
                **defaults
            )
            self.register(inst)

    def process_value_domains(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_VALUE_DOMAIN)

        val_dom = None
        n = 0
        for i, row in enumerate(sheet.iter_rows(row_offset=2)):
            meaning = row[1].value
            value = row[2].value or "-99"
            order = row[2].value or n
            if val_dom is None:
                vd_name = row[0].value
                dt, c = models.DataType.objects.get_or_create(name='Number')
                if c:
                    self.register(dt)
                val_dom, created = models.ValueDomain.objects.get_or_create(
                    name=vd_name,
                    defaults={
                        # "workgroup": wg,
                        "definition": self.DEFAULT_DEFINITION,
                        "data_type": dt,
                    }
                )
                if not created:
                    val_dom.permissiblevalue_set.all().delete()
                self.register(val_dom)
                models.PermissibleValue.objects.create(
                    valueDomain=val_dom,
                    value=value, meaning=meaning, order=order
                )
                n += 1
            elif row[0].value is None and row[1].value is None:
                if n == 1:
                    val_dom.permissiblevalue_set.all().delete()
                val_dom = None
                n = 0
            else:
                models.PermissibleValue.objects.create(
                    valueDomain=val_dom,
                    value=value, meaning=meaning, order=order
                )
                n += 1

    def process_indicators(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_INDICATORS)

        for row in sheet.iter_rows(row_offset=1):
            if row[0].value is None:
                continue
            name = get_col(row, 'A').value

            if not name:
                continue

            defn = get_col(row, 'D')
            if defn:
                defn = defn.value or self.DEFAULT_DEFINITION
            else:
                defn = self.DEFAULT_DEFINITION
            defaults = {
                "name": name,
                "definition": defn,
                "references": get_col(row, 'K').value,
                "rationale": get_col(row, 'P').value if get_col(row, 'M') else "",
                #"short_name": get_col(row,'C').value if get_col(row,'C') else "",
                "comments": get_col(row, 'J').value if get_col(row, 'C') else "",
            #     "disaggregation_description": lb_2_p(row["Disaggregation(s) 1:"])+lb_2_p(row["Disaggregation(s) 2:"]),
                "denominator_description": get_col(row, 'F').value if get_col(row, 'F') else "",
                "numerator_description": get_col(row, 'E').value if get_col(row, 'E') else "",
            }
            ind_identifier = get_col(row, 'B').value
            ind = self.get_from_identifier(ind_identifier)
            if ind:
                comet.Indicator.objects.filter(pk=ind.pk).update(**defaults)
                ind = comet.Indicator.objects.get(pk=ind.pk)
            else:
                ind = comet.Indicator(**defaults)
                ind.save()
                self.register(ind)
                self.make_identifier(ind_identifier, ind)

            des = [de.strip() for de in get_col(row, 'F').value.split(';')]
            for de in des:
                de = models.DataElement.objects.filter(short_name=de).first()
                if de:
                    ind.numerators.add(de)

            instument_name = get_col(row, 'I').value
            instrument = lo_models.Instrument.objects.filter(name=instument_name)
            if instrument.exists():
                instrument.first().indicators.add(ind)
            else:
                print "No instrument", instument_name, instrument

            self.text_to_slots(ind, get_col(row, 'O').value, 'Theory of Change')
            self.text_to_slots(ind, get_col(row, 'Q').value, 'No Poverty')
            self.text_to_slots(ind, get_col(row, 'H').value, 'Data collection method')
            self.text_to_slots(ind, get_col(row, 'C').value, 'Question text', clean=True)
            self.text_to_slots(ind, get_col(row, 'G').value, 'Method of Measurement', clean=True)
            self.text_to_slots(ind, get_col(row, 'N').value, 'Terms of use')
            self.text_to_slots(ind, get_col(row, 'J').value, 'Languages')
            self.text_to_slots(ind, get_col(row, 'L').value, 'Population')
            self.text_to_slots(ind, get_col(row, 'M').value, 'Rationale')

            if get_col(row, 'P'):
                goal_name = get_col(row, 'P').value
                goal = lo_models.Goal.objects.get(short_name=goal_name)
                goal.indicators.add(ind)
                goal.save()

    def process_data_elements(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_DATA_ELEMENTS)
        for row in sheet.iter_rows(row_offset=1):
            if row[1].value is None:
                continue
            name = get_col(row, 'A').value   # [1].value
            vd_name = get_col(row, 'E')
            if vd_name:
                vd = models.ValueDomain.objects.get(name=vd_name.value.strip())
            else:
                vd = None

            defaults = {
                "name": name,
                "valueDomain": vd,  # row["Short name:"],
                "definition": self.DEFAULT_DEFINITION,
            }
            de_identifier = get_col(row, 'B').value
            de = self.get_from_identifier(de_identifier)
            if de:
                models.DataElement.objects.filter(pk=de.pk).update(**defaults)
                de = comet.Indicator.objects.get(pk=de.pk)
            else:
                de = models.DataElement(**defaults)
                de.save()
                self.register(de)
                self.make_identifier(de_identifier, de)

            self.register(de)
            q_name = get_col(row, 'D')
            if q_name:
                q, c = mallard_qr.Question.objects.update_or_create(
                    name=get_col(row, 'D').value.split('?')[0],
                    question_text=get_col(row, 'D').value,
                    defaults={
                        "collected_data_element": de,
                    },
                )
                q.save()
                self.register(q)

            if vd and q_name:
                rd, c = mallard_qr.ResponseDomain.objects.update_or_create(
                    question=q,
                    value_domain=vd,
                )

            ind = self.get_from_identifier(get_col(row, 'C').value)
            if ind:
                ind.numerators.add(de)
                ind.save()

    # Importer Helpers
    def register(self, thing):
        """Aristotle MDR regiter
        """
        models.Status.objects.get_or_create(
            concept=thing,
            registrationAuthority=self.authority,
            registrationDate=self.DEFAULT_REGISTRATION_DATE,
            state=models.STATES.recorded
        )

    def get_from_identifier(self, ident):
        obj = MDR_ID.ScopedIdentifier.objects.filter(
            namespace=self.authority_namespace,
            identifier=ident
        ).first()
        if obj is None:
            return None
        else:
            return obj.concept.item

    def make_identifier(self, ident, item):
        obj = MDR_ID.ScopedIdentifier.objects.create(
            namespace=self.authority_namespace,
            identifier=ident,
            concept=item
        )
        return obj

    def text_to_slots(self, item, col, slot_name, slot_type='', clean=False):
        for val in col.split(';'):
            val = val.strip()
            if clean:
                val = lb_2_p(val, sep="\n")
            if val:
                obj, created = Slot.objects.get_or_create(
                    name=slot_name,
                    type=slot_type,
                    concept=item,
                    value=val
                )
