from __future__ import unicode_literals
from aristotle_mdr import models
from aristotle_mdr.contrib.slots.models import Slot
from comet import models as comet
from datetime import date
from indicators import models as lo_models
from openpyxl import load_workbook
from .utils import BaseImporter, get_col, get_vcol, has_required_cols
from ..models import CategoryOption, Category, CategoryCombination


class IndicatorImporter(BaseImporter):
    """Import indicators from spreadsheet
    """
    DEFAULT_AUTHORITY_NAME = 'Logical Outcomes'
    DEFAULT_AUTHORITY_PREFIX = 'lo'
    DEFAULT_REGISTRATION_DATE = date.today()
    DEFAULT_DEFINITION = 'Imported from the Indicator Reference Sheet'
    # data file sheets
    SHEET_INDICATORS = 'Indicators'
    SHEET_INDICATOR_TYPE = 'Indicator type'
    SHEET_DATA_ELEMENTS = 'Data Elements'
    SHEET_CATEGORY_OPTIONS = 'CategoryOptions'
    SHEET_CATEGORIES = 'Categories'
    SHEET_CATEGORY_COMBOS = 'CategoryCombos'
    SHEET_OPTIONS_SETS = 'OptionSets'

    def __init__(self, data_file, collection=None, status=None):
        self.wb = load_workbook(data_file, read_only=True)
        self.collection = collection or ''
        self.status = status or models.STATES.recorded

    def process(self):
        self.process_authorities()
        self.process_category_options()
        self.process_categories()
        self.process_category_combitation()
        self.process_options_sets()
        self.process_data_elements()
        self.process_indicator_types()
        self.process_indicators()
        # Reload registry
        self.clear_cache()
        self.rebuild_index()

    def process_category_options(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_CATEGORY_OPTIONS)

        for row in sheet.iter_rows(row_offset=1):
            self.log_row(sheet.title, row)

            if not has_required_cols(row, 'A', 'C'):
                continue

            option, c = CategoryOption.objects.update_or_create(
                code=get_col(row, 'C').value,
                collection=self.collection,
                defaults={
                    'name': get_col(row, 'A').value,
                    'short_name': get_col(row, 'D').value,
                }
            )

    def process_categories(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_CATEGORIES)

        categories = []

        for row in sheet.iter_rows(row_offset=1):
            self.log_row(sheet.title, row)

            if not has_required_cols(row, 'A', 'C'):
                continue

            category, c = Category.objects.update_or_create(
                code=get_col(row, 'C').value,
                collection=self.collection,
                defaults={
                    'name': get_col(row, 'A').value,
                    'short_name': get_col(row, 'D').value,
                }
            )

            if category not in categories:
                # remove all options from categories the first time we import it
                categories.append(category)
                category.options.clear()

            option = CategoryOption.objects.get(code=get_col(row, 'E').value)
            category.options.add(option)

    def process_category_combitation(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_CATEGORY_COMBOS)

        category_combinations = []

        for row in sheet.iter_rows(row_offset=1):
            self.log_row(sheet.title, row)

            if not has_required_cols(row, 'A', 'C'):
                continue

            category_combination, c = CategoryCombination.objects.update_or_create(
                code=get_col(row, 'C').value,
                collection=self.collection,
                defaults={
                    'name': get_col(row, 'A').value,
                    'short_name': get_col(row, 'D').value,
                }
            )

            if category_combination not in category_combinations:
                # remove all categories from category_combination the first time we import it
                category_combinations.append(category_combination)
                category_combination.categories.clear()

            category = Category.objects.get(code=get_col(row, 'E').value)
            category_combination.categories.add(category)

    def process_options_sets(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_OPTIONS_SETS)

        # Value domain as Text
        data_type = self.define_data_type('Text')

        value_domains = {}
        for i, row in enumerate(sheet.iter_rows(row_offset=1)):
            self.log_row(sheet.title, row)

            if not has_required_cols(row, 'A', 'C', 'D', 'F'):
                continue

            name = get_col(row, 'A').value
            code = get_col(row, 'C').value
            option_name = get_col(row, 'D').value
            option_code = get_col(row, 'F').value

            if code not in value_domains:
                val_dom = self.get_from_identifier(code)
                if not val_dom:
                    val_dom, created = models.ValueDomain.objects.get_or_create(
                        name=name,
                        definition=self.DEFAULT_DEFINITION,
                        data_type=data_type,
                    )
                    # Add collection as slot field
                    if self.collection:
                        self.text_to_slots(val_dom, self.collection, 'Collection')
                    self.make_identifier(code, val_dom)

                self.register(val_dom)
                val_dom.permissiblevalue_set.all().delete()
                value_domains[code] = {'val_dom': val_dom, 'options': []}

            pv = models.PermissibleValue.objects.create(
                valueDomain=value_domains[code]['val_dom'],
                value=option_code,
                meaning=option_name,
                order=len(value_domains[code]['options']) + 1
            )
            value_domains[code]['options'].append(pv)

    def process_data_elements(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_DATA_ELEMENTS)

        for row in sheet.iter_rows(row_offset=1):
            self.log_row(sheet.title, row)

            if not has_required_cols(row, 'A', 'C'):
                continue

            name = get_vcol(row, 'A')
            code = get_vcol(row, 'C')
            short_name = get_vcol(row, 'D')
            description = get_vcol(row, 'F')
            form_name = get_vcol(row, 'G')
            domain_type = get_vcol(row, 'H')
            value_type = get_vcol(row, 'I')
            aggregation_operator = get_vcol(row, 'J')
            category_combination = get_vcol(row, 'L')
            url = get_vcol(row, 'M')
            zero_is_significant = get_vcol(row, 'N')
            option_set = get_vcol(row, 'O')
            comment_option_set = get_vcol(row, 'P')

            de = self.get_from_identifier(code)
            if not de:
                de, c = models.DataElement.objects.get_or_create(
                    name=name,
                    short_name=short_name,
                    definition=description or name,
                )
                self.make_identifier(code, de)
                # do not remove the slots collection so we can track dependencies
                Slot.objects.filter(concept=de).exclude(name='Collection').delete()

            self.register(de)

            # Add custom properties
            self.text_to_slots(de, form_name, 'Form name')
            self.text_to_slots(de, domain_type, 'Domain type')
            self.text_to_slots(de, value_type, 'Value type')
            self.text_to_slots(de, aggregation_operator, 'Aggregation operator')
            self.text_to_slots(de, category_combination, 'Category combination Code')
            self.text_to_slots(de, zero_is_significant, 'Zero is significant')
            self.text_to_slots(de, comment_option_set, 'Comment option set')
            self.text_to_slots(de, url, 'URL')

            # Add collection as slot field
            if self.collection:
                self.text_to_slots(de, self.collection, 'Collection')

            # Add value domain
            if option_set:
                vd = self.get_from_identifier(option_set)
                if vd:
                    de.valueDomain = vd
                    de.save()
            # get value domain from data type
            # else:

    def process_indicator_types(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_INDICATOR_TYPE)

        for row in sheet.iter_rows(row_offset=1):
            self.log_row(sheet.title, row)

            if not has_required_cols(row, 'A', 'B'):
                continue

            name = get_vcol(row, 'A')
            factor = get_vcol(row, 'B')

            ind_type, c = comet.IndicatorType.objects.update_or_create(
                short_name=name,
                defaults={
                    'name': name,
                    'definition': name,
                }
            )
            self.text_to_slots(ind_type, factor, 'Factor')

            # Add collection as slot field
            if self.collection:
                self.text_to_slots(ind_type, self.collection, 'Collection')

    def process_indicators(self):
        sheet = self.wb.get_sheet_by_name(self.SHEET_INDICATORS)

        for row in sheet.iter_rows(row_offset=2):
            self.log_row(sheet.title, row)

            if not has_required_cols(row, 'A', 'B', 'D'):
                continue

            name = get_vcol(row, 'A')
            short_name = get_vcol(row, 'B')
            code = get_vcol(row, 'D')
            description = get_vcol(row, 'E', default='')
            disaggregation = get_vcol(row, 'F', default='')
            indicator_type = get_vcol(row, 'G')
            numerator = get_vcol(row, 'H')
            numerator_description = get_vcol(row, 'I', default='')
            numerator_computation = get_vcol(row, 'J', default='')
            denominator = get_vcol(row, 'K')
            denominator_description = get_vcol(row, 'L', default='')
            denominator_computation = get_vcol(row, 'M', default='')
            method_of_m = get_vcol(row, 'N')
            instrument_name = get_vcol(row, 'O')
            reference = get_vcol(row, 'P', default='')
            data_source = get_vcol(row, 'Q')
            population = get_vcol(row, 'R')
            rationale = get_vcol(row, 'S', default='')
            terms_of_use = get_vcol(row, 'T')
            languages = get_vcol(row, 'U')
            theory_of_change = get_vcol(row, 'V')
            data_collection = get_vcol(row, 'W')
            sdg = get_vcol(row, 'X')
            outcomes = get_vcol(row, 'Y')
            form_instructions = get_vcol(row, 'Z')

            ind = self.get_from_identifier(code)
            if not ind:
                ind, c = comet.Indicator.objects.update_or_create(
                    short_name=short_name,
                    defaults={
                        'name': name,
                        'definition': description or name,
                        # 'disaggregation_description': disaggregation_description,
                        'numerator_description': numerator_description,
                        'numerator_computation': numerator_computation,
                        'denominator_description': denominator_description,
                        'denominator_computation': denominator_computation,
                        'references': reference,
                        'rationale': rationale,
                    }
                )
                self.make_identifier(code, ind)

            self.register(ind)

            # clean previous data
            Slot.objects.filter(concept=ind).exclude(name='Collection').delete()
            ind.numerators.clear()
            ind.denominators.clear()

            self.results['info']['indicators'].append(ind)

            # Add custom properties as slots
            self.text_to_slots(ind, method_of_m, 'Method of measurement')
            self.text_to_slots(ind, data_source, 'Data source')
            self.text_to_slots(ind, population, 'Population')
            self.text_to_slots(ind, terms_of_use, 'Terms of use')
            self.text_to_slots(ind, disaggregation, 'Disaggregation')
            self.text_to_slots(ind, languages, 'Languages')
            self.text_to_slots(ind, theory_of_change, 'Theory of change')
            self.text_to_slots(ind, data_collection, 'Data collection')
            self.text_to_slots(ind, sdg, 'Sustainable development goals')
            self.text_to_slots(ind, outcomes, 'Outcomes')
            self.text_to_slots(ind, form_instructions, 'Form instructions')

            # Add collection as slot field
            if self.collection:
                self.text_to_slots(ind, self.collection, 'Collection')

            # Add Data Elements
            ind.numerators.add(*self.get_elements(numerator))
            ind.denominators.add(*self.get_elements(denominator))

            # Add insturment relation
            instrument = lo_models.Instrument.objects.filter(name=instrument_name)
            if instrument.exists():
                instrument.first().indicators.add(ind)

            # Add indicator type
            try:
                ind_type = comet.IndicatorType.objects.get(short_name=indicator_type)
                ind.indicatorType = ind_type
                ind.save()
            except comet.IndicatorType.DoesNotExist:
                pass
